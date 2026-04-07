"""
SyllabusCheck — Report Generator Service
Phase 7: Builds PDF (ReportLab) and Excel (openpyxl) gap-analysis reports.
"""
import asyncio
import os
import uuid
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Any

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, inch
from reportlab.platypus import (
    HRFlowable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# Output directory for generated reports
REPORT_DIR = os.getenv("REPORT_DIR", "/app/reports")
os.makedirs(REPORT_DIR, exist_ok=True)

# ── Color constants ──────────────────────────────────────────────────────────

BRAND_ORANGE   = colors.HexColor("#C75B12")
BRAND_DARK     = colors.HexColor("#1C1C1C")
COVERED_GREEN  = colors.HexColor("#10b981")
PARTIAL_AMBER  = colors.HexColor("#f59e0b")
MISSING_RED    = colors.HexColor("#ef4444")
LIGHT_GRAY     = colors.HexColor("#f3f4f6")
MID_GRAY       = colors.HexColor("#9ca3af")
WHITE          = colors.white

# ── Helpers ──────────────────────────────────────────────────────────────────

def _group_by_course(coverage_rows: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in coverage_rows:
        grouped[str(row["course_id"])].append(row)
    return grouped


def _status_color_pdf(status: str):
    return {
        "covered": COVERED_GREEN,
        "partial":  PARTIAL_AMBER,
        "missing":  MISSING_RED,
    }.get(status, MID_GRAY)


# ── PDF Builder ──────────────────────────────────────────────────────────────

async def build_pdf_report(data: dict, report_id: str) -> str:
    """Generate a PDF report and return its file path."""
    path = os.path.join(REPORT_DIR, f"report_{report_id}.pdf")
    await asyncio.to_thread(_build_pdf_sync, data, path)
    return path


def _build_pdf_sync(data: dict, path: str):
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=22,
        textColor=BRAND_DARK,
        spaceAfter=4,
    )
    h2_style = ParagraphStyle(
        "H2",
        parent=styles["Heading2"],
        fontSize=13,
        textColor=BRAND_ORANGE,
        spaceBefore=14,
        spaceAfter=4,
    )
    h3_style = ParagraphStyle(
        "H3",
        parent=styles["Heading3"],
        fontSize=11,
        textColor=BRAND_DARK,
        spaceBefore=10,
        spaceAfter=2,
    )
    body_style = ParagraphStyle(
        "Body",
        parent=styles["Normal"],
        fontSize=9,
        leading=14,
        textColor=colors.HexColor("#374151"),
    )
    small_style = ParagraphStyle(
        "Small",
        parent=styles["Normal"],
        fontSize=8,
        textColor=MID_GRAY,
    )

    doc = SimpleDocTemplate(
        path,
        pagesize=letter,
        topMargin=1.8 * cm,
        bottomMargin=1.8 * cm,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
    )

    story = []
    courses  = data["courses"]
    coverage = data["coverage"]
    grouped  = _group_by_course(coverage)
    summary  = {
        "total_courses": len(courses),
        "covered": sum(1 for r in coverage if r["status"] == "covered"),
        "partial":  sum(1 for r in coverage if r["status"] == "partial"),
        "missing":  sum(1 for r in coverage if r["status"] == "missing"),
        "avg_pct":  round(
            sum(c.get("coverage_score") or 0 for c in courses) / max(len(courses), 1), 1
        ),
    }

    # ── Cover ──────────────────────────────────────────────────────────
    story.append(Spacer(1, 1.5 * cm))
    story.append(Paragraph(data["title"], title_style))
    story.append(Paragraph(
        f"Generated {data['generated_at'][:10]} · {data['user']['name']} ({data['user']['email']})",
        small_style,
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=BRAND_ORANGE, spaceAfter=14))

    # ── Executive summary table ────────────────────────────────────────
    story.append(Paragraph("Executive Summary", h2_style))
    summary_table_data = [
        ["Courses Analysed", "Avg Coverage", "Covered", "Partial", "Missing"],
        [
            str(summary["total_courses"]),
            f"{summary['avg_pct']}%",
            str(summary["covered"]),
            str(summary["partial"]),
            str(summary["missing"]),
        ],
    ]
    t = Table(summary_table_data, colWidths=[3 * cm, 3 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0),  BRAND_DARK),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  WHITE),
        ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0),  9),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT_GRAY, WHITE]),
        ("FONTSIZE",    (0, 1), (-1, -1), 10),
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR",   (1, 1), (1, 1),   BRAND_ORANGE),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("TOPPADDING",  (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5 * cm))

    # ── Per-course sections ───────────────────────────────────────────
    for course in courses:
        cid   = str(course["id"])
        rows  = grouped.get(cid, [])
        score = course.get("coverage_score") or 0

        story.append(Paragraph(
            f"{'📘 ' if score >= 60 else '⚠️ '}{course['title']} "
            f"<font size='8' color='#9ca3af'>{course.get('code','')}</font>",
            h2_style,
        ))

        covered_kws = [r for r in rows if r["status"] == "covered"]
        partial_kws = [r for r in rows if r["status"] == "partial"]
        missing_kws = [r for r in rows if r["status"] == "missing"]

        # Metrics mini-table
        metrics = [
            ["Score", "Domain", "Covered", "Partial", "Missing"],
            [
                f"{score:.0f}%",
                course.get("domain", "—"),
                str(len(covered_kws)),
                str(len(partial_kws)),
                str(len(missing_kws)),
            ],
        ]
        mt = Table(metrics, colWidths=[2.5 * cm, 4 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
        mt.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#374151")),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  WHITE),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 8),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT_GRAY]),
            ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(mt)
        story.append(Spacer(1, 0.3 * cm))

        # Top missing keywords table
        if missing_kws:
            story.append(Paragraph("Top Missing Skills (by job frequency)", h3_style))
            top_missing = sorted(missing_kws, key=lambda r: r.get("frequency", 0), reverse=True)[:15]
            miss_data = [["Skill", "Category", "Job Frequency", "Score"]] + [
                [
                    r["keyword_text"],
                    r.get("category", "—"),
                    str(r.get("frequency", 0)),
                    f"{(r.get('similarity_score', 0) or 0) * 100:.0f}%",
                ]
                for r in top_missing
            ]
            mt2 = Table(miss_data, colWidths=[6 * cm, 3.5 * cm, 3 * cm, 2.5 * cm])
            mt2.setStyle(TableStyle([
                ("BACKGROUND",  (0, 0), (-1, 0),  MISSING_RED),
                ("TEXTCOLOR",   (0, 0), (-1, 0),  WHITE),
                ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
                ("FONTSIZE",    (0, 0), (-1, -1), 8),
                ("ALIGN",       (1, 0), (-1, -1), "CENTER"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#fef2f2"), WHITE]),
                ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#fecaca")),
                ("TOPPADDING",  (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(mt2)
            story.append(Spacer(1, 0.2 * cm))

        # Covered keywords (compact tag-like list)
        if covered_kws:
            story.append(Paragraph("Covered Skills", h3_style))
            kw_text = "  ·  ".join(
                f'<font color="#065f46">{r["keyword_text"]}</font>'
                for r in covered_kws[:30]
            )
            if len(covered_kws) > 30:
                kw_text += f'  <font color="#9ca3af">+{len(covered_kws)-30} more</font>'
            story.append(Paragraph(kw_text, body_style))

        story.append(Spacer(1, 0.5 * cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e5e7eb"), spaceAfter=4))

    doc.build(story)


# ── Excel Builder ────────────────────────────────────────────────────────────

async def build_excel_report(data: dict, report_id: str) -> str:
    """Generate an Excel report and return its file path."""
    path = os.path.join(REPORT_DIR, f"report_{report_id}.xlsx")
    await asyncio.to_thread(_build_excel_sync, data, path)
    return path


def _build_excel_sync(data: dict, path: str):
    wb = openpyxl.Workbook()

    # ── Shared styles ──────────────────────────────────────────────────
    orange_fill = PatternFill("solid", fgColor="C75B12")
    dark_fill   = PatternFill("solid", fgColor="1C1C1C")
    green_fill  = PatternFill("solid", fgColor="D1FAE5")
    amber_fill  = PatternFill("solid", fgColor="FEF3C7")
    red_fill    = PatternFill("solid", fgColor="FEE2E2")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    bold_font   = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    def set_header_row(ws, row_idx: int, headers: list[str], fill=None):
        fill = fill or dark_fill
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=h)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    courses  = data["courses"]
    coverage = data["coverage"]
    grouped  = _group_by_course(coverage)

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 15
    ws_sum.column_dimensions["D"].width = 12
    ws_sum.column_dimensions["E"].width = 12
    ws_sum.column_dimensions["F"].width = 12

    ws_sum["A1"] = data["title"]
    ws_sum["A1"].font = Font(bold=True, size=14, color="1C1C1C")
    ws_sum["A2"] = f"Generated: {data['generated_at'][:10]}  |  {data['user']['name']}"
    ws_sum["A2"].font = Font(size=9, color="6B7280")
    ws_sum.row_dimensions[1].height = 22
    ws_sum.row_dimensions[2].height = 14

    set_header_row(ws_sum, 4, ["Course", "Code", "Domain", "Covered", "Partial", "Missing", "Score"])
    ws_sum.column_dimensions["G"].width = 10

    for row_i, course in enumerate(courses, 5):
        cid   = str(course["id"])
        rows  = grouped.get(cid, [])
        covered = sum(1 for r in rows if r["status"] == "covered")
        partial  = sum(1 for r in rows if r["status"] == "partial")
        missing  = sum(1 for r in rows if r["status"] == "missing")
        score    = course.get("coverage_score") or 0

        vals = [course["title"], course.get("code",""), course.get("domain",""), covered, partial, missing, f"{score:.1f}%"]
        for col, v in enumerate(vals, 1):
            cell = ws_sum.cell(row=row_i, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col == 7:  # score column
                fill_c = green_fill if score >= 60 else (amber_fill if score >= 30 else red_fill)
                cell.fill = fill_c
                cell.font = bold_font

        ws_sum.row_dimensions[row_i].height = 16

    # ── Sheet 2: All Coverage Rows ────────────────────────────────────
    ws_cov = wb.create_sheet("Coverage Detail")
    for col, w in zip("ABCDEFG", [30, 25, 15, 15, 12, 10, 12]):
        ws_cov.column_dimensions[col].width = w

    set_header_row(ws_cov, 1, ["Course", "Skill", "Category", "Domain", "Status", "Score", "Job Freq"])

    for row_i, r in enumerate(coverage, 2):
        course_title = next((c["title"] for c in courses if str(c["id"]) == str(r["course_id"])), "—")
        status       = r["status"]
        score        = (r.get("similarity_score") or 0) * 100
        row_fill = {"covered": green_fill, "partial": amber_fill, "missing": red_fill}.get(status)

        vals = [course_title, r["keyword_text"], r.get("category",""), r.get("domain",""), status, f"{score:.0f}%", r.get("frequency",0)]
        for col, v in enumerate(vals, 1):
            cell = ws_cov.cell(row=row_i, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col == 5 and row_fill:
                cell.fill = row_fill
                cell.font = bold_font

        ws_cov.row_dimensions[row_i].height = 15

    # ── Sheet 3: Top Missing Skills (aggregated) ──────────────────────
    ws_miss = wb.create_sheet("Top Missing Skills")
    for col, w in zip("ABCDE", [30, 20, 12, 12, 35]):
        ws_miss.column_dimensions[col].width = w

    missing_rows = [r for r in coverage if r["status"] == "missing"]
    skill_map: dict[str, dict] = {}
    for r in missing_rows:
        kt = r["keyword_text"]
        if kt not in skill_map:
            skill_map[kt] = {"keyword": kt, "category": r.get("category",""), "frequency": r.get("frequency",0), "courses": []}
        course_title = next((c["title"] for c in courses if str(c["id"]) == str(r["course_id"])), "—")
        skill_map[kt]["courses"].append(course_title)

    sorted_skills = sorted(skill_map.values(), key=lambda x: x["frequency"], reverse=True)

    set_header_row(ws_miss, 1, ["Skill", "Category", "Job Frequency", "Missing From # Courses", "Courses Missing It"])

    for row_i, s in enumerate(sorted_skills, 2):
        vals = [s["keyword"], s["category"], s["frequency"], len(s["courses"]), ", ".join(s["courses"][:5])]
        for col, v in enumerate(vals, 1):
            cell = ws_miss.cell(row=row_i, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 5))
            if col == 3:
                cell.font = Font(bold=True, color="7C3AED")
        ws_miss.row_dimensions[row_i].height = 15

    wb.save(path)
